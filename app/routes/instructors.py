from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from app import db
from app.models import Instructor, InstructorTrack, Track, Profile, HiringPlan
from app.services.scoring_service import calculate_score
from datetime import datetime
from app.services.log_service import log_activity
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask import send_file

instructors_bp = Blueprint('instructors', __name__)

@instructors_bp.route('/')
@login_required
def list_instructors():
    # جلب التراكات عشان نحطها في الفلتر
    tracks = Track.query.all()
    
    # بناء الـ Query الأساسي
    query = Instructor.query
    
    # 1. البحث (Search)
    search_term = request.args.get('q', '').strip()
    if search_term:
        query = query.filter(
            db.or_(
                Instructor.name.ilike(f'%{search_term}%'),
                Instructor.instructor_id.ilike(f'%{search_term}%')
            )
        )
    
    # 2. الفلتر بالتراك (Filter)
    track_filter = request.args.get('track_id', type=int)
    if track_filter:
        # نعمل Subquery عشان نجيب اليوزرز اللي عندهم التراك ده
        subquery = db.session.query(InstructorTrack.instructor_id)\
            .filter(InstructorTrack.track_id == track_filter).subquery()
        query = query.filter(Instructor.id.in_(subquery))
    
    # تنفيذ الـ Query مع إزالة التكرارات (Distinct) عشان لو الانستراكتر عنده التراك ده مرتين
    instructors = query.distinct().order_by(Instructor.created_at.desc()).all()
    
    return render_template('instructors/index.html', 
                           instructors=instructors, 
                           tracks=tracks, 
                           search_term=search_term, 
                           track_filter=track_filter)


@instructors_bp.route('/view/<int:id>')
@login_required
def view_instructor(id):
    inst = Instructor.query.get_or_404(id)
    return render_template('instructors/view.html', instructor=inst)

@instructors_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create_instructor():
    if current_user.role == 'Viewer':
        flash('Viewers cannot add instructors.', 'danger')
        return redirect(url_for('instructors.list_instructors'))
    tracks = Track.query.all()
    plans = HiringPlan.query.all() # جلب الخطط لوضعها في الـ Dropdown
    
    if request.method == 'POST':
        try:
            # تحويل التاريخ
            review_date_str = request.form.get('review_date')
            review_date_obj = datetime.strptime(review_date_str, '%Y-%m-%d').date() if review_date_str else None

            inst = Instructor(
                name=request.form.get('name'),
                instructor_id=request.form.get('instructor_id'),
                phone=request.form.get('phone'),
                cv_url=request.form.get('cv_url'),
                video_url=request.form.get('video_url'),
                rate_per_hour=request.form.get('rate_per_hour') or None,
                employment_type=request.form.get('employment_type'),
                business_type=request.form.get('business_type'),
                technical_feedback=request.form.get('technical_feedback'),
                first_approval=request.form.get('first_approval'),
                second_approval=request.form.get('second_approval'),
                review_date=review_date_obj,
                reviewer_name=request.form.get('reviewer_name'),
                notes=request.form.get('notes'),
                plan_id=request.form.get('plan_id') or None  # ربط الانستراكتر بالخطة
            )
            
            # حساب السكور
            inst.score = calculate_score(inst.first_approval, inst.second_approval, inst.technical_feedback)
            db.session.add(inst)
            db.session.flush()

            # حفظ التراكات والبروفايلات
            for i in range(1, 4):
                track_id = request.form.get(f'track_{i}')
                profile_id = request.form.get(f'profile_{i}')
                if track_id and profile_id:
                    db.session.add(InstructorTrack(
                        instructor_id=inst.id, track_id=track_id, 
                        profile_id=profile_id, order=i
                    ))

            db.session.commit()
            log_activity(action='Create', target_type='Instructor', details=f'Created Instructor: {inst.name}')
            flash('Instructor added successfully!', 'success')
            return redirect(url_for('instructors.list_instructors'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
            
    return render_template('instructors/create.html', tracks=tracks, plans=plans)

@instructors_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_instructor(id):
    if current_user.role != 'Admin':
        flash('Only Admins can edit instructors.', 'danger')
        return redirect(url_for('instructors.list_instructors'))
        
    inst = Instructor.query.get_or_404(id)
    tracks = Track.query.all()
    plans = HiringPlan.query.all() # جلب الخطط لوضعها في الـ Dropdown
    
    # جلب التراكات الحالية للانستراكتر لعرضها في الفورم
    existing_tracks = InstructorTrack.query.filter_by(instructor_id=id).order_by(InstructorTrack.order).all()
    tracks_dict = {t.order: t for t in existing_tracks}
    
    if request.method == 'POST':
        try:
            # تحويل التاريخ
            review_date_str = request.form.get('review_date')
            review_date_obj = datetime.strptime(review_date_str, '%Y-%m-%d').date() if review_date_str else None
            
            # تحديث البيانات الأساسية
            inst.name = request.form.get('name')
            inst.instructor_id = request.form.get('instructor_id')
            inst.phone = request.form.get('phone')
            inst.cv_url = request.form.get('cv_url')
            inst.video_url = request.form.get('video_url')
            inst.rate_per_hour = request.form.get('rate_per_hour') or None
            inst.employment_type = request.form.get('employment_type')
            inst.business_type = request.form.get('business_type')
            inst.technical_feedback = request.form.get('technical_feedback')
            inst.first_approval = request.form.get('first_approval')
            inst.second_approval = request.form.get('second_approval')
            inst.review_date = review_date_obj
            inst.reviewer_name = request.form.get('reviewer_name')
            inst.notes = request.form.get('notes')
            inst.plan_id = request.form.get('plan_id') or None
            inst.status = request.form.get('status', 'New')# تحديث الخطة
            
            # إعادة حساب الـ Score
            inst.score = calculate_score(inst.first_approval, inst.second_approval, inst.technical_feedback)
            
            # حذف التراكات القديمة وإضافة الجديدة
            InstructorTrack.query.filter_by(instructor_id=id).delete()
            
            for i in range(1, 4):
                track_id = request.form.get(f'track_{i}')
                profile_id = request.form.get(f'profile_{i}')
                if track_id and profile_id:
                    db.session.add(InstructorTrack(
                        instructor_id=id, track_id=track_id, 
                        profile_id=profile_id, order=i
                    ))

            db.session.commit()
            log_activity(action='Update', target_type='Instructor', details=f'Updated Instructor: {inst.name}')
            flash('Instructor updated successfully!', 'success')
            return redirect(url_for('instructors.list_instructors'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
            
    return render_template('instructors/edit.html', instructor=inst, tracks=tracks, tracks_dict=tracks_dict, plans=plans)

@instructors_bp.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_instructor(id):
    if current_user.role != 'Admin':
        flash('Only Admins can delete instructors.', 'danger')
        return redirect(url_for('instructors.list_instructors'))
        
    inst = Instructor.query.get_or_404(id)
    db.session.delete(inst)
    db.session.commit()
    log_activity(action='Delete', target_type='Instructor', details=f'Deleted Instructor ID: {id}')
    flash('Instructor deleted.', 'info')
    return redirect(url_for('instructors.list_instructors'))

@instructors_bp.route('/export')
@login_required
def export_instructors():
    # السماح للأدمن فقط بالتصدير
    if current_user.role != 'Admin':
        flash('Only Admins can export data.', 'danger')
        return redirect(url_for('instructors.list_instructors'))
        
    instructors = Instructor.query.order_by(Instructor.created_at.desc()).all()
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructors Data"
    
    # تنسيق الشكل (Styling)
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # كتابة الهيدرات (مطابقة للـ Sheet الأصلي)
    headers = [
        "Instructor Name", "Instructor ID", "Phone Number", 
        "Track 1", "Profile 1", "Track 2", "Profile 2", "Track 3", "Profile 3",
        "CV", "Video", "Rate per Hour (EGP)", 
        "Employment Type", "Business Type", "Technical Feedback",
        "First Approval", "Second Approval", "Review Date", "Reviewer Name",
        "Status", "Hiring Plan", "Notes", "Score"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # كتابة البيانات
    for row_num, inst in enumerate(instructors, 2):
        # استخراج التراكات (نرتبها في Track 1, Track 2, Track 3)
        tracks_data = {1: ("", ""), 2: ("", ""), 3: ("", "")}
        for t in inst.tracks:
            if t.order in tracks_data:
                tracks_data[t.order] = (t.track.name if t.track else "", t.profile.name if t.profile else "")
        
        # استخراج اسم الخطة لو موجود
        plan_name = inst.hiring_plan.title if inst.hiring_plan else ""

        row_data = [
            inst.name,
            inst.instructor_id or "",
            inst.phone or "",
            tracks_data[1][0], tracks_data[1][1],
            tracks_data[2][0], tracks_data[2][1],
            tracks_data[3][0], tracks_data[3][1],
            inst.cv_url or "",
            inst.video_url or "",
            float(inst.rate_per_hour) if inst.rate_per_hour else 0,
            inst.employment_type or "",
            inst.business_type or "",
            inst.technical_feedback or "",
            inst.first_approval or "",
            inst.second_approval or "",
            inst.review_date.strftime('%Y-%m-%d') if inst.review_date else "",
            inst.reviewer_name or "",
            inst.status or "New",
            plan_name,
            inst.notes or "",
            inst.score
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
            
    # تعديل عرض الأعمدة عشان النص يبان كويس
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['U'].width = 20
    ws.column_dimensions['V'].width = 30
    
    # حفظ الملف في الذاكرة (BytesIO) عشان نرسله للمستخدم
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # تسجيل الحدث في الـ Logs
    log_activity(action='Export', target_type='Instructor', details=f'Exported Instructors Data to Excel')
    
    # إرسال الملف للمستخدم للتحميل
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Instructors_Data.xlsx'
    )